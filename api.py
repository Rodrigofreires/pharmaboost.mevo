import asyncio
import base64
import io
import json
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Dict, Any
import pandas as pd
import openpyxl
from pypdf import PdfReader
from bs4 import BeautifulSoup

# Importa os casos de uso da sua aplicação, que contêm a lógica de negócio
from app import use_cases

app = FastAPI(
    title="Gemini Application API",
    description="API para acessar casos de uso baseados no Gemini, com fluxo de revisão humana e otimização contínua.",
    version="4.4.0"  # Versão com Reprocessamento Implementado
)

# --- Configuração do CORS ---
origins = ["http://localhost:5500", "http://127.0.0.1:5500", "null", "*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Modelos Pydantic (sem alterações) ---
class ApprovedItem(BaseModel):
    sku: int
    product_name: str
    html_content: str
    seo_title: str
    meta_description: str

class ReprocessItem(BaseModel):
    sku: int
    product_name: str

class OptimizationRequest(BaseModel):
    product_type: str
    product_name: str
    product_info: Dict[str, Any]

# --- Endpoints da API ---

@app.post("/process-for-review", tags=["Processador de Planilha com Otimização de IA"])
async def process_for_review(
    spreadsheet: UploadFile = File(...),
    bulas: List[UploadFile] = File(...),
    skus_json: str = Form(...)
):
    """
    Recebe uma planilha e múltiplos arquivos de bula para processamento em lote.
    Este endpoint gera e OTIMIZA o conteúdo usando o SeoOptimizerAgent.
    """
    try:
        spreadsheet_bytes = await spreadsheet.read()
        bulas_data = [(bula.filename, await bula.read()) for bula in bulas]
        sku_list = [int(s) for s in json.loads(skus_json)]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler os arquivos: {e}")

    async def event_stream():
        async def send_event(event_type: str, data: dict):
            return f"event: {event_type}\ndata: {json.dumps(data)}\n\n"

        try:
            yield await send_event("log", {"message": "Lendo o arquivo da planilha...", "type": "info"})
            df = pd.read_excel(io.BytesIO(spreadsheet_bytes))
            yield await send_event("log", {"message": "Planilha carregada com sucesso.", "type": "success"})

            if len(sku_list) != len(bulas_data):
                raise ValueError("A quantidade de SKUs não corresponde à de bulas.")

            total_bulas = len(bulas_data)
            yield await send_event("log", {"message": f"Iniciando processamento e otimização de {total_bulas} SKUs...", "type": "info"})

            COLUNA_ID_SKU = "_IDSKU (Não alterável)"
            COLUNA_NOME_PRODUTO = "_NomeProduto (Obrigatório)"
            COLUNA_PALAVRAS_CHAVE = "_PalavrasChave"

            for i, ((bula_filename, bula_bytes), sku) in enumerate(zip(bulas_data, sku_list)):
                progress = f"({i+1}/{total_bulas})"
                log_prefix = f"<b>[SKU: {sku}]</b> {progress}"

                linha_produto = df[df[COLUNA_ID_SKU] == sku]
                if linha_produto.empty:
                    yield await send_event("log", {"message": f"{log_prefix} Não encontrado. Pulando.", "type": "warning"})
                    continue

                nome_produto = linha_produto.iloc[0][COLUNA_NOME_PRODUTO]
                palavras_chave = ""
                if COLUNA_PALAVRAS_CHAVE in linha_produto.columns:
                    palavras_chave = linha_produto.iloc[0][COLUNA_PALAVRAS_CHAVE]
                if pd.isna(palavras_chave) or not palavras_chave:
                    palavras_chave = "bula, para que serve, como usar"

                yield await send_event("log", {"message": f"{log_prefix} Processando '{nome_produto}'...", "type": "info"})

                try:
                    reader = PdfReader(io.BytesIO(bula_bytes))
                    texto_da_bula = "".join(page.extract_text() + "\n" for page in reader.pages)

                    if not texto_da_bula.strip():
                        raise ValueError("Texto do PDF está vazio.")

                    product_info_simulado = {
                        "bula_text": texto_da_bula,
                        "palavras_chave": palavras_chave
                    }

                    yield await send_event("log", {"message": f"{log_prefix} Enviando para o Otimizador com IA...", "type": "info"})

                    optimization_generator = use_cases.run_seo_pipeline_stream(
                        product_type="medicine",
                        product_name=nome_produto,
                        product_info=product_info_simulado
                    )

                    final_content_data = None
                    final_score = 0
                    async for event_chunk in optimization_generator:
                        yield event_chunk

                        if "event: done" in event_chunk:
                            data_str = event_chunk.split('data: ')[1]
                            final_data = json.loads(data_str)
                            final_score = final_data.get("final_score", 0)

                            final_content_data = {
                                "html_content": final_data.get("final_content") or "<p>Erro ao gerar conteúdo.</p>",
                                "seo_title": final_data.get("seo_title") or f"{nome_produto}",
                                "meta_description": final_data.get("meta_description") or "Descrição não gerada."
                            }

                    if final_content_data:
                        review_item = {"sku": sku, "product_name": nome_produto, **final_content_data}
                        yield await send_event("review_item", review_item)

                        if final_score >= 70:
                            yield await send_event("log", {"message": f"{log_prefix} Conteúdo OTIMIZADO (Score Final: {final_score}) gerado. Aguardando sua revisão.", "type": "success"})
                        else:
                            yield await send_event("log", {"message": f"{log_prefix} Melhor score atingido ({final_score}) não alcançou a meta de 70, mas foi enviado para revisão.", "type": "info"})
                    else:
                        yield await send_event("log", {"message": f"{log_prefix} ERRO: O otimizador não retornou um resultado final.", "type": "error"})

                except Exception as e:
                    yield await send_event("log", {"message": f"{log_prefix} ERRO: {e}", "type": "error"})

                await asyncio.sleep(0.5)

        except Exception as e:
            yield await send_event("error", {"message": f"Erro crítico no processamento: {str(e)}", "type": "error"})

    return StreamingResponse(event_stream(), media_type="text/event-stream")

@app.post("/finalize-spreadsheet", tags=["Processador de Planilha com Otimização de IA"])
async def finalize_spreadsheet(
    spreadsheet: UploadFile = File(...),
    approved_data_json: str = Form(...)
):
    """
    Recebe a planilha original e os dados aprovados para montar e retornar o arquivo Excel final.
    """
    try:
        spreadsheet_bytes = await spreadsheet.read()
        approved_data = json.loads(approved_data_json)

        df_for_lookup = pd.read_excel(io.BytesIO(spreadsheet_bytes))
        workbook = openpyxl.load_workbook(io.BytesIO(spreadsheet_bytes))
        sheet = workbook.active

        COLUNA_ID_SKU = "_IDSKU (Não alterável)"
        COLUNA_V_HTML = "_DescricaoProduto"
        COLUNA_AD_TITULO = "_TituloSite"
        COLUNA_AE_META_DESC = "_DescricaoMetaTag"

        header = [cell.value for cell in sheet[1]]
        try:
            sku_col_idx = header.index(COLUNA_ID_SKU) + 1
            html_col_idx = header.index(COLUNA_V_HTML) + 1
            title_col_idx = header.index(COLUNA_AD_TITULO) + 1
            meta_col_idx = header.index(COLUNA_AE_META_DESC) + 1
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Erro Crítico: A coluna obrigatória '{e.args[0]}' não foi encontrada na planilha.")

        for item in approved_data:
            sku = item['sku']
            index_linha_pd = df_for_lookup.index[df_for_lookup[COLUNA_ID_SKU] == sku].tolist()

            if index_linha_pd:
                excel_row_num = index_linha_pd[0] + 2
                sheet.cell(row=excel_row_num, column=html_col_idx).value = item['html_content']
                sheet.cell(row=excel_row_num, column=title_col_idx).value = item['seo_title']
                sheet.cell(row=excel_row_num, column=meta_col_idx).value = item['meta_description']

        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

        return JSONResponse(content={
            "filename": f"planilha_final_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "file_data": excel_base64
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao finalizar a planilha: {str(e)}")

@app.post("/reprocess-items", tags=["Processador de Planilha com Otimização de IA"])
async def reprocess_items(
    spreadsheet: UploadFile = File(...),
    bulas: List[UploadFile] = File(...),
    skus_json: str = Form(...)
):
    """
    Reprocessa uma lista de itens que foram previamente reprovados.
    Esta função reutiliza o fluxo de processamento principal para garantir consistência.
    """
    print("Endpoint de reprocessamento atingido.")
    # A implementação é a mesma que a do endpoint principal, então podemos simplesmente chamar a função dele.
    # O frontend é responsável por enviar apenas os arquivos e SKUs dos itens reprovados.
    return await process_for_review(spreadsheet, bulas, skus_json)

@app.post("/finalize-disapproved-spreadsheet", tags=["Processador de Planilha com Otimização de IA"])
async def finalize_disapproved_spreadsheet(
    spreadsheet: UploadFile = File(...),
    disapproved_data_json: str = Form(...)
):
    """
    Gera uma planilha contendo apenas as linhas dos produtos que foram reprovados.
    """
    try:
        spreadsheet_bytes = await spreadsheet.read()
        df_original = pd.read_excel(io.BytesIO(spreadsheet_bytes))

        disapproved_data = json.loads(disapproved_data_json)
        disapproved_skus = [item['sku'] for item in disapproved_data]
        
        df_disapproved = df_original[df_original['_IDSKU (Não alterável)'].isin(disapproved_skus)].copy()

        output_buffer = io.BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            df_disapproved.to_excel(writer, index=False, sheet_name='Reprovados')
        
        output_buffer.seek(0)
        excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

        return JSONResponse(content={
            "filename": f"planilha_reprovados_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "file_data": excel_base64
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar a planilha de reprovados: {str(e)}")